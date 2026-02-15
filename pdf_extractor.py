#!/usr/bin/env python3
"""
PDF Extraction Script
Extracts invoice data from PDF files (Mogli Labs, SDI, JLL) and outputs
structured data into an Excel file with separate worksheets per company type.
"""

import os
import re
import sys
from datetime import datetime
from pathlib import Path
import argparse

try:
    import PyPDF2
except ImportError:
    print("Error: PyPDF2 is not installed. Please install it using: pip install PyPDF2")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please install it using: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Column definitions per company type (matching the reference Excel format)
# ---------------------------------------------------------------------------

MOGLI_COLUMNS = [
    "Invoice Type", "Vendor Name (Billed From)", "Vendor GSTN (Billed From GSTN)",
    "Buyer Name (Billed To)", "Buyer GSTIN (Billed To GSTN)", "Place of Suply",
    "Invoice No", "Invoice Date", "Description of Goods", "HSN", "Qty", "Unit",
    "Unit Price", "Taxable Total", "IGST    %", "IGST Amount", "CGST %",
    "CGST Amount", "SGST %", "SGST Amount", "Amount",
]

SDI_COLUMNS = [
    "Invoice Type", "Vendor Name (Billed From)", "Vendor GSTN (Billed From GSTN)",
    "Buyer Name (Billed To)", "Buyer GSTIN (Billed To GSTN)", "Invoice No",
    "Invoice Date", "Description of Goods", "HSN/SAC", "Quantity", "Rate", "Per",
    "Taxable Total", "IGST    %", "IGST Amount", "CGST %", "CGST Amount",
    "SGST %", "SGST Amount", "Amount",
]

JLL_COLUMNS = [
    "Invoice Type", "Vendor Name (Billed From)", "Vendor GSTN (Billed From GSTN)",
    "Buyer Name (Billed To)", "Buyer GSTIN (Billed To GSTN)", "Place of Suply",
    "Invoice No", "Invoice Date", "Description", "HSN", "Qty", "Taxable Value",
    "IGST    %", "IGST Amount", "CGST %", "CGST Amount", "SGST %",
    "SGST Amount", "Amount", "Site Name", "Service Type",
]


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _read_pdf_pages(pdf_path):
    """Return list of page texts from a PDF file."""
    pages = []
    with open(pdf_path, "rb") as fh:
        reader = PyPDF2.PdfReader(fh)
        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)
    return pages


def _clean(text):
    """Remove non-breaking spaces and excessive whitespace."""
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def _extract_gstn(text, *labels):
    """Extract GST number following any of the given labels.

    Handles common label variants: GSTIN No, GSTN No, GSTIN, GSTIN/UIN,
    GST NO, GSTNO, etc.
    """
    # Default set of labels when none are provided explicitly
    if not labels:
        labels = ("GSTIN/UIN", "GSTIN No", "GSTN No", "GSTIN", "GST NO", "GSTNO")
    for label in labels:
        pattern = re.compile(
            re.escape(label) + r"\s*:?\s*([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z][A-Z][0-9A-Z])",
            re.IGNORECASE,
        )
        m = pattern.search(text)
        if m:
            return m.group(1)
    return ""


# GSTN regex for extracting all 15-character GST numbers from a block
_GSTN_RE = re.compile(r"\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z][A-Z][0-9A-Z])\b")


def _find_all_unique_gstns(text):
    """Return a list of unique GSTN numbers in order of appearance."""
    seen = []
    for g in _GSTN_RE.findall(text):
        if g not in seen:
            seen.append(g)
    return seen


def _parse_indian_number(s):
    """Parse a number string that may use Indian comma formatting."""
    if not s:
        return 0.0
    s = s.strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _unescape_html(text):
    """Unescape common HTML entities in text."""
    text = text.replace("&lt;", "<")
    text = text.replace("&gt;", ">")
    text = text.replace("&amp;", "&")
    text = text.replace("&quot;", '"')
    text = text.replace("&apos;", "'")
    return text


def _parse_date(date_str):
    """Try to parse a date string into a datetime object."""
    date_str = date_str.strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return date_str


# ---------------------------------------------------------------------------
# Company detection
# ---------------------------------------------------------------------------

COMPANY_PATTERNS = {
    "Mogli Labs": re.compile(r"Mogli\s+Labs", re.IGNORECASE),
    "SDI": re.compile(r"SDI\s+Business\s+Services", re.IGNORECASE),
    "JLL": re.compile(
        r"Jones\s+Lang\s+LaSalle|JLL\s+invoice", re.IGNORECASE
    ),
}


def detect_company(full_text):
    """Detect the company type from the full PDF text.

    Returns one of: 'Mogli Labs', 'SDI', 'JLL', or None.
    """
    for company, pattern in COMPANY_PATTERNS.items():
        if pattern.search(full_text):
            return company
    return None


# ---------------------------------------------------------------------------
# Mogli Labs extractor
# ---------------------------------------------------------------------------

def _extract_mogli(pages):
    """Extract line-item rows from a Mogli Labs invoice PDF."""
    full_text = _clean("\n".join(pages))
    first_page = _clean(pages[0])

    # Invoice type
    invoice_type = "TAX INVOICE"

    # Vendor name: after "Billed From:" / "Name:" or pick from first line
    vendor_name = ""
    m = re.search(r"Billed\s+From\s*:\s*\n?\s*Name\s*:\s*(.+)", first_page, re.IGNORECASE)
    if m:
        vendor_name = m.group(1).strip()
    if not vendor_name:
        # Try first prominent company name line
        m = re.search(r"(Mogli\s+Labs\s*\(?India\)?\s*(?:Private|Pvt\.?)\s+Limited)", full_text, re.IGNORECASE)
        if m:
            vendor_name = m.group(1).strip()

    # Normalize vendor name: remove parentheses and extra spaces, upper-case
    if vendor_name:
        vendor_name = re.sub(r"[()]", "", vendor_name)
        vendor_name = re.sub(r"\s+", " ", vendor_name).strip().upper()

    # Vendor GSTN – try common labels
    vendor_gstn = _extract_gstn(first_page)

    # Buyer name
    buyer_name = ""
    m = re.search(
        r"(?:Detail\s+of\s+)?Buyer\s*\(?Billed\s+To\)?\s*:?\s*\n?\s*Name\s*:\s*(.+)",
        first_page,
        re.IGNORECASE,
    )
    if m:
        buyer_name = m.group(1).strip().upper()

    # Buyer GSTN – look in the "Billed To" section first
    buyer_gstn = ""
    m = re.search(r"(?:Billed\s+To|Buyer)\s*\)?.*?GSTIN?\s*(?:No\.?)?\s*:\s*(\S+)", first_page, re.IGNORECASE | re.DOTALL)
    if m:
        candidate = m.group(1)
        if _GSTN_RE.match(candidate):
            buyer_gstn = candidate
    if not buyer_gstn:
        # Fallback: find all GSTNs and take the second unique one
        unique = _find_all_unique_gstns(first_page)
        if len(unique) >= 2:
            buyer_gstn = unique[1]

    # Place of Supply
    place_of_supply = ""
    m = re.search(r"Place\s+of\s+Supply\s*:\s*(.+)", first_page, re.IGNORECASE)
    if m:
        place_of_supply = m.group(1).strip()

    # Invoice No
    invoice_no = ""
    m = re.search(r"Invoice\s+No\.?\s*:\s*(\S+)", first_page, re.IGNORECASE)
    if m:
        invoice_no = m.group(1).strip()

    # Invoice Date
    invoice_date = ""
    m = re.search(r"Invoice\s+Date\s*:\s*(\S+)", first_page, re.IGNORECASE)
    if m:
        invoice_date = _parse_date(m.group(1).strip())

    # Extract line items from the table
    # Format: S.No Description HSN Qty Unit UnitPrice TaxableTotal CGST% CGSTAmount SGST% SGSTAmount Amount
    rows = []

    # Find the table section – search all pages
    for page_text in pages:
        page_text = _clean(page_text)
        table_start = page_text.find("S.No")
        if table_start < 0:
            table_start = page_text.find("S. No")
        if table_start < 0:
            continue

        grand_total_pos = page_text.find("Grand Total", table_start)
        if grand_total_pos >= 0:
            table_text = page_text[table_start:grand_total_pos]
        else:
            table_text = page_text[table_start:]

        # Remove header line
        header_end = table_text.find("\n")
        if header_end >= 0:
            table_body = table_text[header_end + 1:]
        else:
            table_body = table_text

        # Parse items: each starts with a number at the beginning
        item_pattern = re.compile(
            r"(\d+)"                     # S.No
            r"(.+?)"                     # Description (multi-line captured as group)
            r"(\d{4,10})\s+"             # HSN
            r"([\d.]+)\s*"               # Qty
            r"([A-Z]+(?:\s*\([A-Z]+\))?)\s*"  # Unit
            r"([\d.]+)\s+"               # Unit Price
            r"([\d,]+)\s+"               # Taxable Total
            r"(\d+)\s*%?\s+"             # CGST %
            r"([\d,]+)\s+"               # CGST Amount
            r"(\d+)\s*%?\s+"             # SGST %
            r"([\d,]+)\s+"               # SGST Amount
            r"([\d,]+)",                 # Amount
            re.DOTALL,
        )

        for match in item_pattern.finditer(table_body):
            description = re.sub(r"\s+", " ", match.group(2).strip())
            hsn = match.group(3).strip()
            qty = float(match.group(4))
            unit = re.sub(r"\s+", " ", match.group(5).strip())
            unit_price = float(match.group(6))
            taxable_total = _parse_indian_number(match.group(7))
            cgst_pct = int(match.group(8)) / 100.0
            cgst_amt = _parse_indian_number(match.group(9))
            sgst_pct = int(match.group(10)) / 100.0
            sgst_amt = _parse_indian_number(match.group(11))
            amount = _parse_indian_number(match.group(12))

            # IGST is 0 for intra-state
            igst_pct = 0
            igst_amt = 0

            rows.append({
                "Invoice Type": invoice_type,
                "Vendor Name (Billed From)": vendor_name,
                "Vendor GSTN (Billed From GSTN)": vendor_gstn,
                "Buyer Name (Billed To)": buyer_name,
                "Buyer GSTIN (Billed To GSTN)": buyer_gstn,
                "Place of Suply": place_of_supply,
                "Invoice No": invoice_no,
                "Invoice Date": invoice_date,
                "Description of Goods": _unescape_html(description),
                "HSN": hsn,
                "Qty": qty,
                "Unit": unit,
                "Unit Price": unit_price,
                "Taxable Total": taxable_total,
                "IGST    %": igst_pct,
                "IGST Amount": igst_amt,
                "CGST %": cgst_pct,
                "CGST Amount": cgst_amt,
                "SGST %": sgst_pct,
                "SGST Amount": sgst_amt,
                "Amount": amount,
            })

    return rows


# ---------------------------------------------------------------------------
# SDI extractor
# ---------------------------------------------------------------------------

def _extract_sdi(pages):
    """Extract line-item rows from an SDI Business Services invoice PDF."""
    full_text = _clean("\n".join(pages))
    first_page = _clean(pages[0])

    invoice_type = "TAX INVOICE"

    # Vendor name – extract from the header of the invoice
    vendor_name = ""
    m = re.search(r"(SDI\s+Business\s+Services\s+India\s+Pvt\.?\s+Ltd)", full_text, re.IGNORECASE)
    if m:
        vendor_name = m.group(1).strip()

    # Vendor GSTN – try common label variants
    vendor_gstn = _extract_gstn(first_page)

    # Buyer name – try "Buyer (Bill to)" then "Consignee" as fallback
    buyer_name = ""
    buyer_gstn = ""
    buyer_match = re.search(
        r"Buyer\s*\(Bill\s+to\)\s*\n(.+?)(?:\n|$)", first_page, re.IGNORECASE
    )
    if buyer_match:
        buyer_name = buyer_match.group(1).strip()
    if not buyer_name:
        buyer_match = re.search(
            r"Consignee\s*\(Ship\s+to\)\s*\n(.+?)(?:\n|$)", first_page, re.IGNORECASE
        )
        if buyer_match:
            buyer_name = buyer_match.group(1).strip()

    # Get all GSTNs from text; the first is vendor, subsequent unique ones are buyer
    unique_gstns = _find_all_unique_gstns(first_page)
    if vendor_gstn and len(unique_gstns) >= 2:
        for g in unique_gstns:
            if g != vendor_gstn:
                buyer_gstn = g
                break
    elif len(unique_gstns) >= 2:
        buyer_gstn = unique_gstns[1]

    # Invoice No
    invoice_no = ""
    m = re.search(r"Invoice\s+No\.?\s*(?:e-Way[^\n]*)?\n\s*(\S+)", first_page, re.IGNORECASE)
    if m:
        invoice_no = m.group(1).strip()

    # Invoice Date
    invoice_date = ""
    m = re.search(r"Dated\s*\n\s*(\d{1,2}-[A-Za-z]+-\d{2,4})", first_page, re.IGNORECASE)
    if m:
        invoice_date = _parse_date(m.group(1).strip())

    # Extract line items from all pages
    # SDI format per page: after "Sl Description of Goods Amount Disc. %per Rate Quantity HSN/SAC"
    # Each item: SlNo Description Amount per Rate Quantity unit HSN/SAC
    rows = []
    item_pattern = re.compile(
        r"(\d+)"                          # Sl No
        r"(.+?)"                          # Description (lazy match)
        r"([\d,]+\.\d{2})\s+"             # Amount (total)
        r"(\w+)\s+"                       # Per (unit type)
        r"([\d,]+\.\d{2})\s+"             # Rate
        r"([\d.]+)\s+"                    # Quantity
        r"(\w+)\s+"                       # Unit type again
        r"(\d{5,10})",                    # HSN/SAC
        re.DOTALL,
    )

    for page_text in pages:
        page_text = _clean(page_text)
        # Find the line items section
        header_match = re.search(
            r"Sl\s+Description\s+of\s+Goods\s+Amount.*?HSN/SAC\s*\n\s*No\.\s*\n",
            page_text, re.DOTALL | re.IGNORECASE,
        )
        if not header_match:
            continue

        items_text = page_text[header_match.end():]

        for match in item_pattern.finditer(items_text):
            description = match.group(2).strip()
            description = re.sub(r"\s+", " ", description)
            per = match.group(4).strip()
            rate = _parse_indian_number(match.group(5))
            qty = float(match.group(6))
            hsn = match.group(8).strip()

            taxable_total = rate * qty

            # Default tax rates (will be refined from per-invoice summary)
            cgst_pct = 0.09
            sgst_pct = 0.09
            igst_pct = 0
            igst_amt = 0

            cgst_amt = taxable_total * cgst_pct
            sgst_amt = taxable_total * sgst_pct
            total_amount = taxable_total + igst_amt + cgst_amt + sgst_amt

            rows.append({
                "Invoice Type": invoice_type,
                "Vendor Name (Billed From)": vendor_name,
                "Vendor GSTN (Billed From GSTN)": vendor_gstn,
                "Buyer Name (Billed To)": buyer_name,
                "Buyer GSTIN (Billed To GSTN)": buyer_gstn,
                "Invoice No": invoice_no,
                "Invoice Date": invoice_date,
                "Description of Goods": description,
                "HSN/SAC": hsn,
                "Quantity": qty,
                "Rate": rate,
                "Per": per,
                "Taxable Total": taxable_total,
                "IGST    %": igst_pct,
                "IGST Amount": igst_amt,
                "CGST %": cgst_pct,
                "CGST Amount": cgst_amt,
                "SGST %": sgst_pct,
                "SGST Amount": sgst_amt,
                "Amount": total_amount,
            })

    # Try to extract actual tax rates from the summary table on the last page
    last_page = _clean(pages[-1])
    tax_summary_match = re.search(
        r"HSN/SAC\s+Total.*?Tax\s+Amount.*?Rate.*?Value\s*\n"
        r"(\d{5,10})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*(\d+)\s*%\s+([\d,]+\.\d{2})\s*(\d+)\s*%\s+([\d,]+\.\d{2})",
        last_page,
        re.DOTALL | re.IGNORECASE,
    )
    if tax_summary_match:
        summary_cgst_rate = int(tax_summary_match.group(4)) / 100.0
        summary_sgst_rate = int(tax_summary_match.group(6)) / 100.0 if tax_summary_match.group(6) else summary_cgst_rate
        # Update rows with correct rates from THIS invoice's summary
        for row in rows:
            row["CGST %"] = summary_cgst_rate
            row["SGST %"] = summary_sgst_rate
            taxable = row["Taxable Total"]
            row["CGST Amount"] = taxable * summary_cgst_rate
            row["SGST Amount"] = taxable * summary_sgst_rate
            row["Amount"] = taxable + row["IGST Amount"] + row["CGST Amount"] + row["SGST Amount"]

    return rows


# ---------------------------------------------------------------------------
# JLL extractor
# ---------------------------------------------------------------------------

def _extract_jll(pages):
    """Extract line-item rows from a JLL (Jones Lang LaSalle) invoice PDF."""
    full_text = _clean("\n".join(pages))
    first_page = _clean(pages[0])

    invoice_type = "TAX INVOICE"

    # Vendor name – extract from the "Bill From" section
    vendor_name = ""
    m = re.search(r"Bill\s+From\s*:\s*\n\s*(.+?)(?:\n|$)", first_page, re.IGNORECASE)
    if m:
        vendor_name = _clean(m.group(1))
    if not vendor_name:
        # Fallback: look for the known company pattern anywhere in the text
        m = re.search(
            r"(Jones\s+Lang\s+LaSalle\s+Property\s+Consultants\s*\(?India\)?\s*(?:Private|Pvt\.?)\s+Ltd\.?)",
            full_text,
            re.IGNORECASE,
        )
        if m:
            vendor_name = m.group(1).strip()

    # Vendor GSTN – try common label variants
    vendor_gstn = _extract_gstn(first_page)

    # Buyer name – extract from "Bill To" section
    buyer_name = ""
    m = re.search(r"Bill\s+To\s*:\s*\n\s*(.+?)(?:\n|$)", first_page, re.IGNORECASE)
    if m:
        raw_name = _clean(m.group(1))
        # Fix joined words from PDF text extraction (e.g. "ElectricITBusiness" → "Electric IT Business")
        raw_name = re.sub(r"([a-z])([A-Z])", r"\1 \2", raw_name)
        # Fix spacing in common abbreviations
        raw_name = re.sub(r"\bI\s*T\s*Business\b", "IT Business", raw_name)
        raw_name = re.sub(r"\bIndia\s+Private\s*$", "India Private Limited", raw_name)
        buyer_name = raw_name

    # If buyer name looks incomplete (fewer than 10 chars likely means truncated
    # or only a fragment was captured), try the summary page entity name as fallback
    if not buyer_name or len(buyer_name) < 10:
        summary_page = _clean(pages[-1]) if pages else ""
        m = re.search(r"Entity\s+Name\s*.*?\n\s*\S+\s+(.+?)\s+SE-IN", summary_page, re.IGNORECASE)
        if m:
            buyer_name = _clean(m.group(1))

    # Buyer GSTN
    buyer_gstn = ""
    unique_gstns = _find_all_unique_gstns(first_page)
    if vendor_gstn and len(unique_gstns) >= 2:
        for g in unique_gstns:
            if g != vendor_gstn:
                buyer_gstn = g
                break
    elif len(unique_gstns) >= 2:
        buyer_gstn = unique_gstns[1]

    # Place of Supply – extract and strip QR-code junk characters that may follow
    place_of_supply = ""
    m = re.search(r"Place\s+of\s+Supply\s*:\s*(\w[\w\s]*)", first_page, re.IGNORECASE)
    if m:
        raw_pos = m.group(1).strip()
        # QR code data often appears as long uppercase strings right after the state name
        # without any space.  Split on a lowercase→uppercase(5+) boundary to remove junk.
        raw_pos = re.split(r"(?<=[a-z])(?=[A-Z]{5,})", raw_pos)[0]
        # Also strip any remaining long uppercase blocks that are clearly not state names
        raw_pos = re.split(r"\s+[A-Z]{5,}", raw_pos)[0]
        place_of_supply = raw_pos.strip()

    # Invoice Number
    invoice_no = ""
    m = re.search(r"Invoice\s+Number\s*:\s*(\S+)", first_page, re.IGNORECASE)
    if m:
        invoice_no = m.group(1).strip()

    # Invoice Date
    invoice_date = ""
    m = re.search(r"Invoice\s+Date\s*:\s*(\d{1,2}-[A-Za-z]+-\d{2,4})", first_page, re.IGNORECASE)
    if m:
        invoice_date = _parse_date(m.group(1).strip())

    # Extract summary table from the last page (page 3 typically)
    # Format: JLL invoice No | Entity Name | Site Name | Service Type | Service Month | HSN | ...
    summary_page = _clean(pages[-1]) if pages else ""

    site_name = ""
    service_type_from_summary = ""
    hsn_from_summary = ""

    summary_header_match = re.search(
        r"JLL\s+invoice\s+No\s+Entity\s+Name\s+Site\s+Name\s+Service\s+Type\s+Service\s+Month\s+HSN",
        summary_page,
        re.IGNORECASE,
    )

    summary_data_lines = []
    if summary_header_match:
        after_header = summary_page[summary_header_match.end():]
        # Parse lines until "Grand Total"
        for line in after_header.split("\n"):
            line = line.strip()
            if not line or line.startswith("Grand Total"):
                break
            summary_data_lines.append(line)

    # Parse the summary data line to get site_name, service_type, hsn
    if summary_data_lines:
        summary_line = summary_data_lines[0]
        # Extract: invoice_no entity_name site_name service_type service_month hsn ...
        # The summary line starts with the invoice number.
        # Service Type may be multi-word (e.g. "M&E Services", "Pantry Services",
        # "Soft Services", "Housekeeping Services")
        m = re.match(
            r"(\S+)\s+"                                      # JLL invoice No
            r"(.+?)\s+"                                      # Entity Name
            r"(SE-IN\s+[^\s]+(?:\s*[-&]\s*[^\s]+)*(?:\s*\([^)]*\))?)\s+"  # Site Name
            r"((?:M&E|Pantry|Soft|Hard|Housekeeping)\s*\w*(?:\s+\w+)?)\s+"  # Service Type
            r"(\d{2}-\d{2}-\d{4})\s+"                        # Service Month
            r"(\d{5,10})",                                   # HSN
            summary_line,
            re.IGNORECASE,
        )
        if m:
            site_name = m.group(3).strip()
            service_type_from_summary = m.group(4).strip()
            hsn_from_summary = m.group(6).strip()

    # Extract line items from invoice pages (pages 1 & 2 typically)
    # Look for the section after "Description HSN Qty UOM Taxable Value"
    rows = []

    # Combine all pages except last (summary) for line item extraction
    invoice_pages_text = _clean("\n".join(pages[:-1])) if len(pages) > 1 else first_page

    # Find line items section - may appear on multiple pages
    # Collect all text after "Description HSN Qty" headers
    desc_sections = []
    for page_text in (pages[:-1] if len(pages) > 1 else pages):
        pt = _clean(page_text)
        idx = pt.find("Description HSN Qty")
        if idx >= 0:
            # Skip the header line itself
            header_end = pt.find("\n", idx)
            if header_end >= 0:
                section = pt[header_end + 1:]
            else:
                section = pt[idx:]
            desc_sections.append(section)

    combined_items_text = "\n".join(desc_sections)

    if combined_items_text:
        # Each line item block ends with Subtotal + CGST + SGST + total value
        # Split by Subtotal to isolate items, but also need the CGST/SGST info
        # that appears after the Subtotal of each item
        # Strategy: find each HSN occurrence with its taxable value, then work
        # backwards to get the description

        # Pattern to find: description text ... HSN_CODE TAXABLE_VALUE
        # followed by Subtotal TAXABLE_VALUE, CGST%, SGST%
        hsn_pattern = re.compile(
            r"(.+?)"                       # Description block
            r"\b(\d{6})\s+"                # HSN code (6 digits)
            r"([\d,]+\.?\d*)\s*\n"         # Taxable value
            r"\s*Subtotal\s+([\d,]+\.?\d*)" # Subtotal value
        , re.DOTALL)

        current_site = site_name
        current_service = service_type_from_summary

        for match in hsn_pattern.finditer(combined_items_text):
            desc_block = match.group(1).strip()
            hsn = match.group(2)
            taxable_value = _parse_indian_number(match.group(3))

            # Clean description: remove noise lines, page numbers, etc.
            desc_block = re.sub(r"^Page\s+\d+.*?\n", "", desc_block, flags=re.IGNORECASE)
            # Remove previous item's CGST/SGST/total lines that might be captured
            desc_block = re.sub(r"Karnataka\s+[CS]GST\d+%\s+[\d,.]+\s*\n?", "", desc_block)
            desc_block = re.sub(r"^[\d,.]+\s*\n", "", desc_block)  # Remove standalone numbers (totals)
            desc_block = re.sub(r"DSC\s+Expiry.*?\n", "", desc_block, flags=re.IGNORECASE)
            desc_block = re.sub(r"Total\s+Amount\s+Due.*?\n", "", desc_block, flags=re.IGNORECASE)
            desc_block = re.sub(r"DS\s+JONES.*?\n", "", desc_block, flags=re.IGNORECASE)

            # Extract site name if present
            site_match = re.search(r"(SE-IN\s+[^\n]+?\([^)]*\))", desc_block, re.IGNORECASE)
            if site_match:
                current_site = site_match.group(1).strip()

            # Clean and join description lines
            desc_lines = [l.strip() for l in desc_block.split("\n") if l.strip()]
            description = " ".join(desc_lines)

            # Look for CGST/SGST after Subtotal in the remaining text
            after_match_start = match.end()
            after_text = combined_items_text[after_match_start:after_match_start + 200]

            cgst_pct = 0.09
            sgst_pct = 0.09
            igst_pct = 0
            igst_amt = 0

            cgst_match = re.search(r"CGST\s*(\d+)\s*%\s+([\d,]+\.?\d*)", after_text)
            sgst_match = re.search(r"SGST\s*(\d+)\s*%\s+([\d,]+\.?\d*)", after_text)
            if cgst_match:
                cgst_pct = int(cgst_match.group(1)) / 100.0
            if sgst_match:
                sgst_pct = int(sgst_match.group(1)) / 100.0

            cgst_amt = taxable_value * cgst_pct
            sgst_amt = taxable_value * sgst_pct
            total_amount = taxable_value + igst_amt + cgst_amt + sgst_amt

            # Determine service type from description keywords or summary
            service_type = ""
            if re.search(r"food\s+service|pantry|catering|Other\s+contract\s+food", description, re.IGNORECASE):
                service_type = "Pantry Services"
            elif re.search(r"M\s*&\s*[ER]|maintenance|mechanical|electrical", description, re.IGNORECASE):
                service_type = "M&E Services"
            elif re.search(r"housekeeping|cleaning", description, re.IGNORECASE):
                service_type = "Housekeeping Services"
            elif re.search(r"soft\s+service", description, re.IGNORECASE):
                service_type = "Soft Services"

            if not service_type:
                service_type = current_service

            rows.append({
                "Invoice Type": invoice_type,
                "Vendor Name (Billed From)": vendor_name,
                "Vendor GSTN (Billed From GSTN)": vendor_gstn,
                "Buyer Name (Billed To)": buyer_name,
                "Buyer GSTIN (Billed To GSTN)": buyer_gstn,
                "Place of Suply": place_of_supply,
                "Invoice No": invoice_no,
                "Invoice Date": invoice_date,
                "Description": description,
                "HSN": hsn,
                "Qty": 0,
                "Taxable Value": taxable_value,
                "IGST    %": igst_pct,
                "IGST Amount": igst_amt,
                "CGST %": cgst_pct,
                "CGST Amount": cgst_amt,
                "SGST %": sgst_pct,
                "SGST Amount": sgst_amt,
                "Amount": total_amount,
                "Site Name": current_site,
                "Service Type": service_type,
            })

    return rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _write_excel(all_data, output_path):
    """Write extracted data to an Excel workbook with separate sheets per company."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    sheets_created = False

    company_configs = {
        "Mogli Labs": ("Mogli Lab ", MOGLI_COLUMNS),
        "SDI": ("SDI ", SDI_COLUMNS),
        "JLL": ("JLL", JLL_COLUMNS),
    }

    for company_type, (sheet_name, columns) in company_configs.items():
        company_rows = all_data.get(company_type, [])
        if not company_rows:
            continue

        ws = wb.create_sheet(title=sheet_name)
        sheets_created = True

        # Write header row
        if company_type == "JLL":
            # JLL has a merged header "Summary Sheet" at T1 then headers at row 2
            ws.cell(row=1, column=20, value="Summary Sheet")
            header_row = 2
        else:
            header_row = 1

        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=header_row, column=col_idx, value=col_name)

        # Write data rows
        data_start_row = header_row + 1
        for row_idx, row_data in enumerate(company_rows, start=data_start_row):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

    if not sheets_created:
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value="No invoice data could be extracted from the provided PDFs.")
    else:
        # Remove default empty sheet
        if default_sheet.title == "Sheet" and default_sheet.max_row <= 1:
            wb.remove(default_sheet)

    wb.save(output_path)
    print(f"✓ Excel output saved to: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process_pdf(pdf_path):
    """Process a single PDF and return (company_type, rows)."""
    try:
        pages = _read_pdf_pages(pdf_path)
    except Exception as e:
        print(f"Warning: Could not read {pdf_path}: {e}")
        return None, []

    full_text = "\n".join(pages)
    company = detect_company(full_text)
    if not company:
        print(f"Warning: Could not identify company type in {pdf_path}")
        return None, []

    print(f"  Detected: {company} in {os.path.basename(pdf_path)}")

    extractors = {
        "Mogli Labs": _extract_mogli,
        "SDI": _extract_sdi,
        "JLL": _extract_jll,
    }

    extractor = extractors.get(company)
    if not extractor:
        return None, []

    try:
        rows = extractor(pages)
    except Exception as e:
        print(f"Warning: Error extracting data from {pdf_path}: {e}")
        return company, []

    return company, rows


def main():
    """Main function to orchestrate PDF extraction."""
    parser = argparse.ArgumentParser(
        description="Extract invoice data from PDF files and output as Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pdf_extractor.py -i /path/to/input/folder -o /path/to/output
  python pdf_extractor.py --input-dir ./input --output-dir ./output
        """,
    )

    parser.add_argument(
        "-i", "--input-dir",
        required=True,
        help="Path to the input folder containing PDF files",
    )

    parser.add_argument(
        "-o", "--output-dir",
        required=True,
        help="Directory to save the Excel output (will be created if it does not exist)",
    )

    args = parser.parse_args()

    # Validate input directory
    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        print(f"Error: Input directory not found: {args.input_dir}")
        sys.exit(1)
    if not input_dir.is_dir():
        print(f"Error: Path is not a directory: {args.input_dir}")
        sys.exit(1)

    # Collect PDF files
    pdf_files = sorted(
        [f for f in input_dir.iterdir() if f.suffix.lower() == ".pdf"]
    )
    if not pdf_files:
        print(f"Error: No PDF files found in {args.input_dir}")
        sys.exit(1)

    # Create output directory
    output_dir = Path(args.output_dir)
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"Error creating output directory: {e}")
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF file(s) in {input_dir}")

    # Process each PDF
    all_data = {}
    for pdf_path in pdf_files:
        print(f"Processing: {pdf_path.name}")
        company, rows = process_pdf(str(pdf_path))
        if company and rows:
            all_data.setdefault(company, []).extend(rows)
            print(f"  Extracted {len(rows)} row(s)")
        elif company:
            print(f"  No line items extracted")

    # Write Excel output
    output_file = output_dir / "extracted_invoices.xlsx"
    _write_excel(all_data, str(output_file))

    # Summary
    total_rows = sum(len(r) for r in all_data.values())
    print(f"\n✓ Extraction complete: {total_rows} total row(s) from {len(all_data)} company type(s)")


if __name__ == "__main__":
    main()
